from tablespam.Excel._as_excel.create_test_files import create_test_files_cars
from tablespam.Excel._as_excel.locations import Locations
import openpyxl
import polars as pl


def test_excel(tmp_path):
    test_xlsx = create_test_files_cars()

    # The created xlsx files correspond to the excel files in data.
    # We load those files and compare our results in test_xlsx to them
    for tst in test_xlsx.wbs:
        # because openpyxl may not read the workbook the same it was written, we
        # will first first save our new workbook and then read it again.
        test_xlsx.wbs[tst].save(f'{tmp_path}/tmp_file.xlsx')
        to_test = openpyxl.load_workbook(filename=f'{tmp_path}/tmp_file.xlsx')
        # load excel file from data
        target = openpyxl.load_workbook(filename=f'tests/data/{tst}.xlsx')
        # Get sheet dimensions
        rows = range(target['Table'].min_row, target['Table'].max_row + 1)
        cols = range(target['Table'].min_column, target['Table'].max_column + 1)
        # Compare the newly saved table with the reference table:
        identical = True
        for row in rows:
            for col in cols:
                if not identical:
                    raise ValueError('Mismatch between test files and excel workbooks.')
                cell_address = f'{openpyxl.utils.get_column_interval(col, col)[0]}{row}'
                # compare values
                if target['Table'][cell_address].value is None:
                    if to_test['Table'][cell_address].value is None:
                        next
                    else:
                        raise ValueError(
                            'Mismatch between test files and excel workbooks.'
                        )
                identical = identical & (
                    target['Table'][cell_address].value
                    == to_test['Table'][cell_address].value
                )
                if not identical:
                    raise ValueError('Mismatch between test files and excel workbooks.')
                # compare styles
                identical = identical & (
                    target['Table'][cell_address].style
                    == to_test['Table'][cell_address].style
                )
                if not identical:
                    raise ValueError('Mismatch between test files and excel workbooks.')
                # borders
                identical = identical & (
                    (
                        target['Table'][cell_address].border.left,
                        target['Table'][cell_address].border.right,
                        target['Table'][cell_address].border.top,
                        target['Table'][cell_address].border.bottom,
                    )
                    == (
                        to_test['Table'][cell_address].border.left,
                        to_test['Table'][cell_address].border.right,
                        to_test['Table'][cell_address].border.top,
                        to_test['Table'][cell_address].border.bottom,
                    )
                )
                if not identical:
                    raise ValueError('Mismatch between test files and excel workbooks.')

                identical = identical & (
                    (
                        target['Table'][cell_address].border.left,
                        target['Table'][cell_address].border.right,
                        target['Table'][cell_address].border.top,
                        target['Table'][cell_address].border.bottom,
                    )
                    == (
                        to_test['Table'][cell_address].border.left,
                        to_test['Table'][cell_address].border.right,
                        to_test['Table'][cell_address].border.top,
                        to_test['Table'][cell_address].border.bottom,
                    )
                )

                if not identical:
                    raise ValueError('Mismatch between test files and excel workbooks.')

                identical = identical & (
                    (
                        target['Table'][cell_address].font.bold,
                        target['Table'][cell_address].font.size,
                    )
                    == (
                        to_test['Table'][cell_address].font.bold,
                        to_test['Table'][cell_address].font.size,
                    )
                )
                if not identical:
                    raise ValueError('Mismatch between test files and excel workbooks.')

        # Additionally, we will check if the data was written correctly
        if tst == 'cars_offset':
            start_row = 3
            start_col = 5
        else:
            start_row = 1
            start_col = 1
        locations = Locations(
            tbl=test_xlsx.tbls[tst], start_row=start_row, start_col=start_col
        )

        read_data = pl.read_excel(
            f'{tmp_path}/tmp_file.xlsx',
            engine='calamine',
            read_options={
                'header_row': locations.get_row('end_row_header') - 1,
                'n_rows': locations.get_row('end_row_data')
                - locations.get_row('end_row_header'),
            },
        )
        if test_xlsx.tbls[tst].table_data['row_data'] is not None:
            expected_data = pl.concat(
                [
                    test_xlsx.tbls[tst].table_data['row_data'],
                    test_xlsx.tbls[tst].table_data['col_data'],
                ],
                how='horizontal',
            )
            row_name_cols = list(
                range(
                    0,
                    locations.get_col('end_col_header_lhs')
                    - locations.get_col('start_col_header_lhs')
                    + 1,
                )
            )
        else:
            expected_data = test_xlsx.tbls[tst].table_data['col_data']

        for row in range(0, expected_data.shape[0]):
            for col in range(0, expected_data.shape[1]):
                # for the headers, we have to take into account that some of the
                # rows may be merged in the row names
                if (read_data[row, col] is None) and (
                    expected_data[row, col] is not None
                ):
                    if (col in row_name_cols) and (
                        expected_data[row - 1, col] == expected_data[row, col]
                    ):
                        next
                    else:
                        raise ValueError('Mismatch between expected and read data.')
                elif (read_data[row, col] is None) and (
                    expected_data[row, col] is None
                ):
                    next
                elif abs(expected_data[row, col] - read_data[row, col]) < 1e-6:
                    next
                else:
                    raise ValueError('Mismatch between expected and read data.')
